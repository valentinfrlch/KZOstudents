from ast import Add
from matplotlib.pyplot import fill
import openpyxl
import geopy.geocoders
import progressbar
from geopy.exc import GeocoderTimedOut, GeocoderUnavailable

path = 'dataset/students.xlsx'


def get_addresses(path):
    # load excel file
    wb = openpyxl.load_workbook(path)
    # get sheet
    sheet = wb.get_sheet_by_name('Sheet1')

    #load data into list
    data = []
    for row in range(2, sheet.max_row + 1):
        try:
            student = {}
            student['klasse'] = sheet['A' + str(row)].value.strip()
            student['name'] = sheet['B' + str(row)].value.strip()
            student['vorname'] = sheet['C' + str(row)].value.strip()
            student['addresse'] = sheet['D' +
                                        str(row)].value.strip() + ' ' + sheet['E' + str(row)].value.strip()
            student['telefon'] = sheet['F' + str(row)].value
            student['email'] = sheet['G' + str(row)].value.strip()
        except (ValueError, AttributeError) as e:
            pass
        data.append(student)

    #get  all addresses into one list
    addresses = []
    for student in data:
        addresses.append(student['addresse'])

    return addresses


def get_coordinates(address):
    geolocator = geopy.geocoders.Nominatim(
        user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36")
    location = geolocator.geocode(address)
    return location.latitude, location.longitude


def visualize_coordinates(coordinates):
    import folium
    map = folium.Map(location=coordinates, zoom_start=12)
    for coordinate in coordinates:
        folium.Marker(coordinate, popup='Ausgezeichnet').add_to(map)
    map.save('map.html')
    import webbrowser
    webbrowser.open('map.html')


def main():
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('Sheet1')
    # add lat and long to the sheet
    for row in range(2, sheet.max_row + 1):
        # if sheet['H' + str(row)].value is empty
        if sheet['H' + str(row)].value is None:
            try:
                coordinates = get_coordinates(
                    sheet['D' + str(row)].value.strip() + ' ' + sheet['E' + str(row)].value.strip())
                sheet['H' + str(row)].value = coordinates[0]
                sheet['I' + str(row)].value = coordinates[1]
            except (AttributeError, ValueError, GeocoderTimedOut, GeocoderUnavailable) as e:
                pass
        # save the file every 100 rows
        if row % 100 == 0:
            wb.save(path)
    wb.save(path)
